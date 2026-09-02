#!/usr/bin/env python3
"""
Build the Exam Proctor Duty site from "Proctor Schedule by Date.xlsx".

Reads the `Exam` sheet (header on row 5, data from row 6) and the `Proctors`
sheet (roster, names in column A from row 6), then writes:

    data.json    the extracted schedule, for reference / reuse
    index.html   the self-contained site (template.html + data baked in)

Only exams on or after WINDOW_START are included.

Usage:
    py -3 build.py                       # use the default source path
    py -3 build.py "path\\to\\Proctor Schedule by Date.xlsx"

Environment overrides:
    PROCTOR_XLSX     source workbook path
    WINDOW_START     earliest exam date to include, as YYYY-MM-DD
"""

import datetime as dt
import json
import os
import re
import shutil
import sys
import tempfile
import warnings

warnings.filterwarnings("ignore")  # openpyxl warns about unsupported extensions

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is missing. Install it with:  py -3 -m pip install openpyxl")

# --- configuration -----------------------------------------------------------

SOURCE_XLSX = os.environ.get("PROCTOR_XLSX") or (
    r"C:\Users\Faris.Alsalem\OneDrive - AL-Hussien bin Abdullah Technical University"
    r"\Faris HTU\Exam Committee\Proctor Schedule by Date.xlsx"
)

# The committee only tracks the final exam period; earlier sittings are done.
WINDOW_START = os.environ.get("WINDOW_START", "2026-09-03")

HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(HERE, "template.html")
OUT_JSON = os.path.join(HERE, "data.json")
OUT_HTML = os.path.join(HERE, "index.html")
OUT_FRAG = os.path.join(HERE, "artifact.html")  # same page, no <html>/<head> wrapper

HEADER_ROW = 5
FIRST_DATA_ROW = 6
PROCTOR_COLS = range(10, 20)  # 0-based: "Proctor 1" .. "Proctor 10"

# --- cell helpers ------------------------------------------------------------


def text(v):
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt.time):
        return v.strftime("%H:%M")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def number(v):
    return int(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def clean_time(v):
    """Times arrive either as real times or as strings like '09:30' / '9:30'."""
    s = text(v)
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    return "%02d:%s" % (int(m.group(1)), m.group(2)) if m else s


def slug(s):
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


# --- extraction --------------------------------------------------------------


def load_roster(wb):
    if "Proctors" not in wb.sheetnames:
        return []
    names = []
    for row in wb["Proctors"].iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        name = text(row[0])
        if name and name.lower() not in ("name", "total"):
            names.append(name)
    return names


def load_exams(wb):
    exams, skipped = [], 0
    for idx, row in enumerate(
        wb["Exam"].iter_rows(min_row=FIRST_DATA_ROW, values_only=True), start=FIRST_DATA_ROW
    ):
        course = text(row[0])
        if not course or course.upper() == "TOTAL":
            continue

        date = text(row[3])
        if date and date < WINDOW_START:
            skipped += 1
            continue

        proctors = [p for p in (text(row[c]) for c in PROCTOR_COLS) if p]
        required = number(row[9])
        assigned = len(proctors)

        # Recompute the shortfall from the assignment cells. The workbook's own
        # "Still Needed" column cannot be trusted: row 22 points at row 25, and
        # rows 23, 24, 33 and 35 have no formula at all.
        shortfall = max(0, required - assigned) if required is not None else None

        start, end = clean_time(row[5]), clean_time(row[6])
        exams.append(
            {
                "id": "%s-%s-%s" % (date or "nodate", slug(course), start or "notime"),
                "sourceRow": idx,
                "course": course,
                "coordinator": text(row[1]),
                "dept": text(row[2]),
                "date": date,
                "day": text(row[4]),
                "start": start,
                "end": end,
                "time": ("%s - %s" % (start, end)) if start and end else (start or end or "TBC"),
                "room": text(row[7]) or "TBC",
                "students": number(row[8]),
                "required": required,
                "assigned": assigned,
                "shortfall": shortfall,
                "proctors": proctors,
                "notes": text(row[22]),
            }
        )

    exams.sort(key=lambda e: (e["date"] or "9999", e["start"] or "99:99", e["course"]))
    return exams, skipped


def proctor_index(exams, roster):
    """Everyone on the roster gets an entry, even with zero duties in the window."""
    index = {name: [] for name in roster}
    for e in exams:
        for name in e["proctors"]:
            index.setdefault(name, []).append(e["id"])
    return [
        {"name": n, "duties": d, "count": len(d), "onRoster": n in roster}
        for n, d in sorted(index.items(), key=lambda kv: kv[0].lower())
    ]


def course_index(exams):
    """One entry per course title — a course can sit more than one exam."""
    index = {}
    for e in exams:
        c = index.setdefault(
            e["course"],
            {"course": e["course"], "coordinator": e["coordinator"], "dept": e["dept"], "sittings": []},
        )
        c["sittings"].append(e["id"])
        if e["coordinator"]:
            c["coordinator"] = e["coordinator"]
        if e["dept"]:
            c["dept"] = e["dept"]
    return sorted(index.values(), key=lambda c: c["course"].lower())


# --- output ------------------------------------------------------------------


def render_body(payload):
    """template.html with the data baked in — no document wrapper."""
    with open(TEMPLATE, encoding="utf-8") as fh:
        tpl = fh.read()
    if "/*__DATA__*/ null" not in tpl:
        sys.exit("template.html is missing the '/*__DATA__*/ null' placeholder.")

    # </script> inside the JSON would close the host <script> element early.
    blob = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    return tpl.replace("/*__DATA__*/ null", blob)


def render_html(body):
    return (
        "<!doctype html>\n"
        '<html lang="en">\n<head>\n'
        '<meta charset="utf-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
        '<meta name="description" content="Proctor duty schedule for the HTU School of '
        'Engineering and Technology final exam period.">\n'
        # The page names staff and says where they will be. It is hosted on a
        # public GitHub Pages site, so keep it out of search results: nobody
        # should reach it by googling a colleague's name. robots.txt says the
        # same thing for crawlers that check it first.
        '<meta name="robots" content="noindex, nofollow, noarchive, nosnippet">\n'
        '<meta name="googlebot" content="noindex, nofollow">\n'
        "<style>img{max-width:100%}[hidden]{display:none!important}</style>\n"
        + body
        + "\n</body>\n</html>\n"
    )


def open_workbook(src):
    """Load the workbook via a temporary copy.

    Excel and LibreOffice hold the file open with sharing denied, so reading it
    directly raises PermissionError while anyone has it open. Copying is still
    permitted, so copy first and read the copy — this lets the sync run whether
    or not the sheet happens to be open.
    """
    tmp = os.path.join(tempfile.gettempdir(), "_proctor_sync_%d.xlsx" % os.getpid())
    try:
        try:
            shutil.copy2(src, tmp)
        except (PermissionError, OSError) as err:
            sys.exit(
                "Could not read the workbook:\n  %s\n\n%s\n\n"
                "If it is open, save and close it, then run this again." % (src, err)
            )
        return openpyxl.load_workbook(tmp, data_only=True)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else SOURCE_XLSX
    if not os.path.exists(src):
        sys.exit("Source workbook not found:\n  %s" % src)

    wb = open_workbook(src)
    exams, skipped = load_exams(wb)
    roster = load_roster(wb)
    short = [e for e in exams if e["shortfall"]]

    payload = {
        "generatedAt": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "sourceFile": os.path.basename(src),
        "sourceModified": dt.datetime.fromtimestamp(os.path.getmtime(src))
        .astimezone()
        .isoformat(timespec="seconds"),
        "windowStart": WINDOW_START,
        "summary": {
            "exams": len(exams),
            "totalStudents": sum(e["students"] or 0 for e in exams),
            "seatsRequired": sum(e["required"] or 0 for e in exams),
            "seatsAssigned": sum(e["assigned"] for e in exams),
            "examsShort": len(short),
            "seatsShort": sum(e["shortfall"] or 0 for e in short),
            "examsWithoutTarget": len([e for e in exams if e["required"] is None]),
            "proctors": len({p for e in exams for p in e["proctors"]}),
            "rosterSize": len(roster),
        },
        "exams": exams,
        "proctorIndex": proctor_index(exams, roster),
        "courseIndex": course_index(exams),
    }

    body = render_body(payload)
    with open(OUT_JSON, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=1, ensure_ascii=False)
    with open(OUT_HTML, "w", encoding="utf-8") as fh:
        fh.write(render_html(body))
    with open(OUT_FRAG, "w", encoding="utf-8") as fh:
        fh.write(body)

    s = payload["summary"]
    print("Source     : %s" % src)
    print("Window     : %s onwards (%d earlier sitting(s) excluded)" % (WINDOW_START, skipped))
    print("Exams      : %d, %d students" % (s["exams"], s["totalStudents"]))
    print("Proctors   : %d / %d seats filled, %d people used of %d on roster"
          % (s["seatsAssigned"], s["seatsRequired"], s["proctors"], s["rosterSize"]))
    if short:
        print("SHORT      : %d exam(s), %d seat(s) open" % (len(short), s["seatsShort"]))
        for e in short:
            print("   %s %s  %s  (needs %d more)" % (e["date"], e["time"], e["course"], e["shortfall"]))
    else:
        print("SHORT      : none - every exam is fully proctored")
    if s["examsWithoutTarget"]:
        print("NOTE       : %d exam(s) have no '# of Proctors' target set" % s["examsWithoutTarget"])
    print("Wrote      : data.json, index.html, artifact.html")


if __name__ == "__main__":
    main()
