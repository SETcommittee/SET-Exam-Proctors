#!/usr/bin/env python3
"""
Build the macro-enabled reminder workbook.

Takes the committee's "Proctor Schedule by Date.xlsx", adds two sheets
(Emails, Reminders) and the VBA behind them, and saves the result alongside
the original as .xlsm. The original .xlsx is never modified.
"""

import os
import re
import shutil
import sys
import tempfile
import warnings

warnings.filterwarnings("ignore")

import openpyxl
import win32com.client as win32
from win32com.client import constants as c  # noqa: F401  (late-bound anyway)

FOLDER = (
    r"C:\Users\Faris.Alsalem\OneDrive - AL-Hussien bin Abdullah Technical University"
    r"\Faris HTU\Exam Committee"
)
SRC_XLSX = os.path.join(FOLDER, "Proctor Schedule by Date.xlsx")
OUT_XLSM = os.path.join(FOLDER, "Proctor Schedule by Date.xlsm")

HERE = os.path.dirname(os.path.abspath(__file__))
BAS_MODULE = os.path.join(HERE, "vba_reminders.bas")
BAS_SHEET = os.path.join(HERE, "vba_sheet.bas")

XL_STD_MODULE = 1
XL_OPENXML_MACRO = 52  # xlOpenXMLWorkbookMacroEnabled

# Names that are the same person written differently. Used only to warn the
# user on the Emails sheet - the macro still matches on the exact string.
ALIAS_HINTS = [
    ("Rajaie Nassar", "Rajaie Ghassan Fawzi Nassar", "Rajaee Nassaer"),
    ("Nezar AlQudah", "Nezar Yahya Abdallah Alqudah"),
    ("Lina AlSamameh", "Lena Ali Salem Al-Samameh"),
    ("Ala BaniIssa", "Ala' Abdel Rahman Sanad Bani Issa"),
    ("Nouralhuda Alenaizat", "Nouralhuda Yousef Jamil Alenaizat"),
    ("Ahmad Samha", "Ahmad Yousef Kamal Samha"),
    ("Mohammad Raed", "Mohammad Ra`ed Mustafa Khirfan"),
]


def txt(v):
    return "" if v is None else str(v).strip()


def suggested_email(name):
    """firstname.lastname@htu.edu.jo - a starting point, never authoritative."""
    cleaned = re.sub(r"[^A-Za-z \-']", " ", name)
    parts = [p for p in cleaned.split() if len(p) > 1]
    if len(parts) < 2:
        return ""
    first = parts[0].lower()
    last = parts[-1].lower().replace("-", "").replace("'", "")
    return "%s.%s@htu.edu.jo" % (first, last)


def collect_people(path):
    tmp = os.path.join(tempfile.gettempdir(), "_collect.xlsx")
    shutil.copy2(path, tmp)
    wb = openpyxl.load_workbook(tmp, data_only=True)

    roster, coords, used = [], [], set()
    for row in wb["Proctors"].iter_rows(min_row=6, values_only=True):
        n = txt(row[0])
        if n and n.lower() not in ("name", "total"):
            roster.append(n)

    for row in wb["Exam"].iter_rows(min_row=6, values_only=True):
        if not txt(row[0]) or txt(row[0]).upper() == "TOTAL":
            continue
        co = txt(row[1])
        if co and co not in coords:
            coords.append(co)
        for cell in row[10:20]:
            if txt(cell):
                used.add(txt(cell))

    os.remove(tmp)

    people = []
    for n in sorted(set(roster) | set(coords) | used, key=str.lower):
        is_p = n in roster or n in used
        is_c = n in coords
        role = "Proctor & coordinator" if (is_p and is_c) else ("Coordinator" if is_c else "Proctor")
        note = ""
        for group in ALIAS_HINTS:
            if n in group:
                others = [x for x in group if x != n]
                note = "Same person as: " + "; ".join(others)
                break
        people.append((n, role, suggested_email(n), note))
    return people


def main():
    if not os.path.exists(SRC_XLSX):
        sys.exit("Source workbook not found:\n  %s" % SRC_XLSX)

    people = collect_people(SRC_XLSX)
    print("Collected %d people." % len(people))

    # Work on a private copy so an open Excel window cannot interfere.
    staging = os.path.join(tempfile.gettempdir(), "_proctor_build.xlsx")
    shutil.copy2(SRC_XLSX, staging)

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        wb = excel.Workbooks.Open(staging)

        for stale in ("Emails", "Reminders"):
            for sh in list(wb.Worksheets):
                if sh.Name == stale:
                    sh.Delete()

        build_emails_sheet(wb, people)
        build_reminders_sheet(wb)
        inject_vba(wb)

        if os.path.exists(OUT_XLSM):
            os.remove(OUT_XLSM)
        wb.SaveAs(OUT_XLSM, FileFormat=XL_OPENXML_MACRO)
        wb.Close(SaveChanges=False)
        print("Wrote %s" % OUT_XLSM)
    finally:
        excel.Quit()
        if os.path.exists(staging):
            try:
                os.remove(staging)
            except OSError:
                pass


def build_emails_sheet(wb, people):
    ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
    ws.Name = "Emails"

    ws.Range("A1").Value = "Email addresses"
    ws.Range("A1").Font.Size = 15
    ws.Range("A1").Font.Bold = True

    ws.Range("A2").Value = (
        "Put each person's real address in the Email column. The reminder macro reads "
        "only that column. Column D is a guess to save typing - check it before copying across."
    )
    ws.Range("A2").Font.Color = 0x8A7266

    for i, head in enumerate(["Name", "Role", "Email", "Suggested (verify!)", "Note"], start=1):
        cell = ws.Cells(3, i)
        cell.Value = head
        cell.Font.Bold = True
        cell.Interior.Color = 0xF6F1E1
        cell.Borders(9).LineStyle = 1  # xlEdgeBottom

    for r, (name, role, guess, note) in enumerate(people, start=4):
        ws.Cells(r, 1).Value = name
        ws.Cells(r, 2).Value = role
        ws.Cells(r, 3).Interior.Color = 0xF2FAF2  # the column to fill in
        ws.Cells(r, 4).Value = guess
        ws.Cells(r, 4).Font.Color = 0x9A9A9A
        if note:
            ws.Cells(r, 5).Value = note
            ws.Cells(r, 5).Font.Color = 0x2A5BA6

    ws.Columns("A").ColumnWidth = 34
    ws.Columns("B").ColumnWidth = 21
    ws.Columns("C").ColumnWidth = 32
    ws.Columns("D").ColumnWidth = 32
    ws.Columns("E").ColumnWidth = 46
    ws.Rows(3).Font.Bold = True
    ws.Activate()
    excel_freeze(wb, "A4")


def build_reminders_sheet(wb):
    ws = wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count))
    ws.Name = "Reminders"

    ws.Range("A1").Value = "Exam reminders"
    ws.Range("A1").Font.Size = 15
    ws.Range("A1").Font.Bold = True
    ws.Range("A2").Value = (
        "Click any exam row to preview its email below. Double-click its Send cell, "
        "or use the buttons, to send."
    )
    ws.Range("A2").Font.Color = 0x8A7266

    ws.Range("A4").Value = "TO"
    ws.Range("A5").Value = "SUBJECT"
    ws.Range("A6").Value = "BODY"
    for row in (4, 5, 6):
        ws.Cells(row, 1).Font.Bold = True
        ws.Cells(row, 1).VerticalAlignment = -4160  # xlTop
        ws.Cells(row, 1).Font.Color = 0x8A7266

    for rng in ("B4:K4", "B5:K5", "B6:K6"):
        ws.Range(rng).Merge()
        ws.Range(rng).Interior.Color = 0xFAF8F5
        ws.Range(rng).VerticalAlignment = -4160
        ws.Range(rng).WrapText = True
    ws.Range("B6").RowHeight = 150
    ws.Rows(4).RowHeight = 30
    ws.Rows(5).RowHeight = 18

    headers = ["Date", "Day", "Time", "Course", "Room", "Coordinator",
               "Proctors", "Recipients", "No address for", "When", "Send", "Last sent"]
    for i, head in enumerate(headers, start=1):
        cell = ws.Cells(9, i)
        cell.Value = head
        cell.Font.Bold = True
        cell.Interior.Color = 0xF6F1E1
        cell.Borders(9).LineStyle = 1

    widths = [12, 10, 14, 40, 26, 30, 44, 40, 26, 11, 9, 16]
    for i, w in enumerate(widths, start=1):
        ws.Columns(i).ColumnWidth = w

    add_button(ws, 210, 118, 118, 24, "Refresh list", "RefreshReminders")
    add_button(ws, 336, 118, 168, 24, "Send for selected row", "SendSelectedReminder")
    add_button(ws, 512, 118, 150, 24, "Preview in Outlook", "PreviewSelectedReminder")
    add_button(ws, 670, 118, 150, 24, "Send all upcoming", "SendAllUpcoming")

    excel_freeze(wb, "A10")


def add_button(ws, left, top, width, height, caption, macro):
    btn = ws.Buttons().Add(left, top, width, height)
    btn.Caption = caption
    btn.OnAction = macro
    btn.Font.Size = 10


def excel_freeze(wb, cell):
    ws = wb.ActiveSheet
    wb.Application.ActiveWindow.FreezePanes = False
    ws.Range(cell).Select()
    wb.Application.ActiveWindow.FreezePanes = True


def inject_vba(wb):
    proj = wb.VBProject

    with open(BAS_MODULE, encoding="utf-8") as fh:
        module_code = fh.read()
    comp = proj.VBComponents.Add(XL_STD_MODULE)
    comp.Name = "mReminders"
    comp.CodeModule.AddFromString(module_code)

    with open(BAS_SHEET, encoding="utf-8") as fh:
        sheet_code = fh.read()
    target = wb.Worksheets("Reminders")
    proj.VBComponents(target.CodeName).CodeModule.AddFromString(sheet_code)

    print("VBA injected: mReminders + Reminders sheet events")


if __name__ == "__main__":
    main()
